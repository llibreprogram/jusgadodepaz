"""
Script para generar plantillas de importación en formato CSV y Excel
con casos de ejemplo para el Sistema de Gestión de Casos
"""

import pandas as pd
from datetime import datetime
import os

def generar_plantillas():
    """Genera plantillas CSV y Excel con casos de ejemplo"""
    
    # Definir las columnas en el orden correcto
    columnas = [
        'numero_carpeta',
        'categoria',
        'etapa_procesal',
        'victima',
        'investigado',
        'fecha_denuncia',
        'fecha_formalizacion',
        'fecha_acusacion',
        'fecha_sentencia',
        'fecha_archivo',
        'monto_reparacion',
        'estado_actual',
        'resultado',
        'apelacion',
        'fiscal_asignado',
        'tiene_citacion',
        'fecha_emision_citacion',
        'fecha_comparecencia',
        'estado_citacion',
        'observaciones_citacion',
        'tiene_orden_arresto',
        'fecha_emision_orden',
        'estado_orden',
        'fecha_cumplimiento_orden',
        'observaciones_orden',
        'origen_orden_arresto',
        'fiscal_inicial',
        'departamento_actual',
        'fiscal_cierre'
    ]
    
    # Casos de ejemplo con todos los campos
    casos_ejemplo = [
        {
            'numero_carpeta': '2024-001',
            'categoria': 'Violencia doméstica',
            'etapa_procesal': 'Investigación',
            'victima': 'María González',
            'investigado': 'Juan Pérez',
            'fecha_denuncia': '2024-01-15',
            'fecha_formalizacion': '2024-01-20',
            'fecha_acusacion': '',
            'fecha_sentencia': '',
            'fecha_archivo': '',
            'monto_reparacion': '5000.00',
            'estado_actual': 'Activo',
            'resultado': 'Pendiente',
            'apelacion': '0',
            'fiscal_asignado': 'Fiscal Ana Martínez',
            'tiene_citacion': '1',
            'fecha_emision_citacion': '2024-01-18',
            'fecha_comparecencia': '2024-02-05',
            'estado_citacion': 'Pendiente',
            'observaciones_citacion': 'Primera citación',
            'tiene_orden_arresto': '0',
            'fecha_emision_orden': '',
            'estado_orden': '',
            'fecha_cumplimiento_orden': '',
            'observaciones_orden': '',
            'origen_orden_arresto': '',
            'fiscal_inicial': 'Fiscal Ana Martínez',
            'departamento_actual': 'San Salvador',
            'fiscal_cierre': ''
        },
        {
            'numero_carpeta': '2024-002',
            'categoria': 'Hurto',
            'etapa_procesal': 'Juicio',
            'victima': 'Comercial El Sol',
            'investigado': 'Carlos Ramírez',
            'fecha_denuncia': '2024-02-01',
            'fecha_formalizacion': '2024-02-05',
            'fecha_acusacion': '2024-02-20',
            'fecha_sentencia': '',
            'fecha_archivo': '',
            'monto_reparacion': '2500.00',
            'estado_actual': 'En juicio',
            'resultado': 'Pendiente',
            'apelacion': '0',
            'fiscal_asignado': 'Fiscal Roberto López',
            'tiene_citacion': '1',
            'fecha_emision_citacion': '2024-02-03',
            'fecha_comparecencia': '2024-02-15',
            'estado_citacion': 'Compareció',
            'observaciones_citacion': 'Asistió a la audiencia',
            'tiene_orden_arresto': '0',
            'fecha_emision_orden': '',
            'estado_orden': '',
            'fecha_cumplimiento_orden': '',
            'observaciones_orden': '',
            'origen_orden_arresto': '',
            'fiscal_inicial': 'Fiscal Roberto López',
            'departamento_actual': 'Santa Ana',
            'fiscal_cierre': ''
        },
        {
            'numero_carpeta': '2024-003',
            'categoria': 'Estafa',
            'etapa_procesal': 'Sentencia',
            'victima': 'Pedro Hernández',
            'investigado': 'Luis Morales',
            'fecha_denuncia': '2023-10-15',
            'fecha_formalizacion': '2023-10-20',
            'fecha_acusacion': '2023-11-10',
            'fecha_sentencia': '2024-01-30',
            'fecha_archivo': '',
            'monto_reparacion': '15000.00',
            'estado_actual': 'Cerrado',
            'resultado': 'Culpable',
            'apelacion': '0',
            'fiscal_asignado': 'Fiscal Carmen Flores',
            'tiene_citacion': '1',
            'fecha_emision_citacion': '2023-10-18',
            'fecha_comparecencia': '2023-11-05',
            'estado_citacion': 'No compareció',
            'observaciones_citacion': 'No se presentó a la audiencia',
            'tiene_orden_arresto': '1',
            'fecha_emision_orden': '2023-11-06',
            'estado_orden': 'Cumplida',
            'fecha_cumplimiento_orden': '2023-11-15',
            'observaciones_orden': 'Capturado por la PNC',
            'origen_orden_arresto': 'Por no comparecencia a cita',
            'fiscal_inicial': 'Fiscal Carmen Flores',
            'departamento_actual': 'La Libertad',
            'fiscal_cierre': 'Fiscal Carmen Flores'
        }
    ]
    
    # Crear DataFrame
    df = pd.DataFrame(casos_ejemplo, columns=columnas)
    
    # Crear directorio para plantillas si no existe
    os.makedirs('plantillas', exist_ok=True)
    
    # Generar CSV
    csv_filename = 'plantillas/plantilla_importacion_casos.csv'
    df.to_csv(csv_filename, index=False, encoding='utf-8-sig')
    print(f"✅ Plantilla CSV creada: {csv_filename}")
    
    # Generar Excel con formato
    excel_filename = 'plantillas/plantilla_importacion_casos.xlsx'
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Casos', index=False)
        
        # Obtener el workbook y worksheet para formatear
        workbook = writer.book
        worksheet = writer.sheets['Casos']
        
        # Ajustar anchos de columna
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Formatear encabezados en negrita
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Agregar hoja de instrucciones
        ws_instrucciones = workbook.create_sheet('Instrucciones', 0)
        
        instrucciones = [
            ['INSTRUCCIONES PARA IMPORTACIÓN DE CASOS'],
            [''],
            ['1. FORMATO DE FECHAS'],
            ['   Todas las fechas deben estar en formato: YYYY-MM-DD (ejemplo: 2024-01-15)'],
            ['   Dejar en blanco si no tiene fecha'],
            [''],
            ['2. CATEGORÍAS VÁLIDAS'],
            ['   - Violencia doméstica'],
            ['   - Hurto'],
            ['   - Robo'],
            ['   - Estafa'],
            ['   - Lesiones'],
            ['   - Amenazas'],
            ['   - Daños'],
            ['   - Apropiación indebida'],
            ['   - Extorsión'],
            ['   - Secuestro'],
            ['   - Homicidio'],
            ['   - Violación'],
            ['   - Tráfico de drogas'],
            ['   - Corrupción'],
            ['   - Fraude'],
            ['   - Otros'],
            [''],
            ['3. ETAPAS PROCESALES VÁLIDAS'],
            ['   - Denuncia'],
            ['   - Investigación'],
            ['   - Formalización'],
            ['   - Audiencia preliminar'],
            ['   - Juicio'],
            ['   - Sentencia'],
            ['   - Apelación'],
            ['   - Ejecución'],
            ['   - Archivo'],
            [''],
            ['4. ESTADOS ACTUALES VÁLIDOS'],
            ['   - Activo'],
            ['   - En juicio'],
            ['   - Cerrado'],
            ['   - Archivado'],
            [''],
            ['5. RESULTADOS VÁLIDOS'],
            ['   - Pendiente'],
            ['   - Culpable'],
            ['   - Inocente'],
            ['   - Desestimado'],
            ['   - Archivado'],
            [''],
            ['6. ESTADOS DE CITACIÓN VÁLIDOS'],
            ['   - Pendiente'],
            ['   - Compareció'],
            ['   - No compareció'],
            ['   - Cancelada'],
            [''],
            ['7. ESTADOS DE ORDEN DE ARRESTO VÁLIDOS'],
            ['   - Pendiente de cumplimiento'],
            ['   - Cumplida'],
            ['   - Cancelada'],
            ['   - Revocada'],
            [''],
            ['8. ORIGEN DE ORDEN DE ARRESTO VÁLIDO'],
            ['   - Directa con denuncia'],
            ['   - Por no comparecencia a cita'],
            ['   - Orden judicial posterior'],
            ['   - Otro'],
            [''],
            ['9. DEPARTAMENTOS VÁLIDOS'],
            ['   - San Salvador'],
            ['   - Santa Ana'],
            ['   - San Miguel'],
            ['   - La Libertad'],
            ['   - Sonsonate'],
            ['   - Usulután'],
            ['   - La Paz'],
            ['   - Ahuachapán'],
            ['   - Morazán'],
            ['   - La Unión'],
            ['   - Chalatenango'],
            ['   - Cuscatlán'],
            ['   - Cabañas'],
            ['   - San Vicente'],
            [''],
            ['10. CAMPOS BOOLEANOS (tiene_citacion, tiene_orden_arresto, apelacion)'],
            ['    Use: 1 para Sí, 0 para No'],
            [''],
            ['11. CAMPOS OBLIGATORIOS'],
            ['    - numero_carpeta (debe ser único)'],
            ['    - categoria'],
            ['    - etapa_procesal'],
            ['    - victima'],
            ['    - investigado'],
            ['    - fiscal_asignado'],
            [''],
            ['12. NOTAS IMPORTANTES'],
            ['    - No modifique los nombres de las columnas'],
            ['    - Elimine esta hoja antes de importar'],
            ['    - Los casos de ejemplo en la hoja "Casos" son una guía'],
            ['    - Puede agregar tantas filas como necesite'],
            ['    - Todos los archivos se guardan en la carpeta "exports/"']
        ]
        
        for row_idx, row_data in enumerate(instrucciones, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_instrucciones.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                elif 'VÁLIDAS' in str(value) or 'VÁLIDOS' in str(value):
                    cell.font = Font(bold=True, size=11)
                elif str(value).startswith('   -'):
                    cell.font = Font(italic=True)
        
        # Ajustar ancho de columna de instrucciones
        ws_instrucciones.column_dimensions['A'].width = 80
    
    print(f"✅ Plantilla Excel creada: {excel_filename}")
    print(f"\n📁 Las plantillas se guardaron en la carpeta 'plantillas/'")
    print(f"\n📋 Cada plantilla contiene 3 casos de ejemplo para guiarte.")
    print(f"   Puedes modificarlos o agregar más filas según necesites.")

if __name__ == '__main__':
    print("=" * 60)
    print("  GENERADOR DE PLANTILLAS DE IMPORTACIÓN")
    print("  Sistema de Gestión de Casos v3.0")
    print("=" * 60)
    print()
    generar_plantillas()
    print("\n✅ Proceso completado exitosamente!")
