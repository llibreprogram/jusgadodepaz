import pandas as pd
from datetime import datetime
import os

class ExportService:
    """Enhanced export service with better formatting and options"""
    
    def __init__(self):
        self.export_dir = 'exports'
        self._ensure_export_dir()
    
    def _ensure_export_dir(self):
        """Create exports directory if it doesn't exist"""
        os.makedirs(self.export_dir, exist_ok=True)
    
    def export_to_csv(self, cases, filename=None, filtered=False, single=False):
        """Export cases to CSV with automatic filename generation"""
        df = pd.DataFrame([vars(c) for c in cases])
        
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            if single and len(cases) == 1:
                # Use case number in filename for single exports
                case_num = cases[0].numero_carpeta.replace('/', '_').replace('\\', '_')
                filename = f'caso_{case_num}_{timestamp}.csv'
            else:
                prefix = 'filtrado' if filtered else 'completo'
                filename = f'casos_{prefix}_{timestamp}.csv'
        
        filepath = os.path.join(self.export_dir, filename)
        df.to_csv(filepath, index=False, encoding='utf-8-sig')
        return filepath
    
    def export_to_excel(self, cases, filename=None, filtered=False, single=False):
        """Export cases to Excel with formatting"""
        df = pd.DataFrame([vars(c) for c in cases])
        
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            if single and len(cases) == 1:
                # Use case number in filename for single exports
                case_num = cases[0].numero_carpeta.replace('/', '_').replace('\\', '_')
                filename = f'caso_{case_num}_{timestamp}.xlsx'
            else:
                prefix = 'filtrado' if filtered else 'completo'
                filename = f'casos_{prefix}_{timestamp}.xlsx'
        
        filepath = os.path.join(self.export_dir, filename)
        
        # Create Excel with formatting
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Casos', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Casos']
            
            # Auto-adjust column width
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format header
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color='22c55e', end_color='22c55e', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        return filepath
    
    def export_statistics_report(self, stats, filename=None):
        """Export statistical summary to Excel"""
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'estadisticas_{timestamp}.xlsx'
        
        filepath = os.path.join(self.export_dir, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = {
                'Métrica': ['Casos Resueltos', 'Casos Pendientes', 'Tasa de Éxito (%)', 
                           'Tiempo Promedio Resolución (días)', 'Casos con Apelación',
                           'Porcentaje Apelaciones (%)'],
                'Valor': [
                    stats.get('resolved', 0),
                    stats.get('pending', 0),
                    round(stats.get('success_rate', 0), 2),
                    round(stats.get('avg_resolution_time', 0), 2),
                    stats.get('appeal_count', 0),
                    round(stats.get('appeal_percentage', 0), 2)
                ]
            }
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Resumen', index=False)
            
            # Cases by category
            if stats.get('cases_by_category'):
                cat_df = pd.DataFrame(list(stats['cases_by_category'].items()), 
                                     columns=['Categoría', 'Cantidad'])
                cat_df = cat_df.sort_values('Cantidad', ascending=False)
                cat_df.to_excel(writer, sheet_name='Por Categoría', index=False)
            
            # Cases by state
            if stats.get('cases_by_estado'):
                estado_df = pd.DataFrame(list(stats['cases_by_estado'].items()), 
                                        columns=['Estado', 'Cantidad'])
                estado_df = estado_df.sort_values('Cantidad', ascending=False)
                estado_df.to_excel(writer, sheet_name='Por Estado', index=False)
            
            # Cases by fiscal
            if stats.get('by_judge'):
                fiscal_df = pd.DataFrame(list(stats['by_judge'].items()), 
                                        columns=['Fiscal', 'Casos Asignados'])
                fiscal_df = fiscal_df.sort_values('Casos Asignados', ascending=False)
                fiscal_df.to_excel(writer, sheet_name='Por Fiscal', index=False)
        
        return filepath

    def export_complete_statistics(self, cases, filename=None):
        """Export comprehensive statistics with multiple sheets grouped by area"""
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'estadisticas_completas_{timestamp}.xlsx'
        
        filepath = os.path.join(self.export_dir, filename)
        
        # Group cases by area
        pension_cases = []
        transito_cases = []
        otros_cases = []
        
        for case in cases:
            cat_lower = case.categoria.lower() if case.categoria else ''
            if 'pensión' in cat_lower or 'pension' in cat_lower:
                pension_cases.append(case)
            elif 'tránsito' in cat_lower or 'transito' in cat_lower:
                transito_cases.append(case)
            else:
                otros_cases.append(case)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # 1. RESUMEN GENERAL
            summary_data = {
                'Área': ['Pensión Alimentaria', 'Tránsito', 'Otros Casos', 'TOTAL'],
                'Cantidad': [
                    len(pension_cases),
                    len(transito_cases),
                    len(otros_cases),
                    len(cases)
                ],
                'Porcentaje': [
                    f"{(len(pension_cases)/len(cases)*100):.1f}%" if cases else "0%",
                    f"{(len(transito_cases)/len(cases)*100):.1f}%" if cases else "0%",
                    f"{(len(otros_cases)/len(cases)*100):.1f}%" if cases else "0%",
                    "100.0%"
                ]
            }
            
            # Add total pension amount if applicable
            total_pension = sum(float(c.monto_pension or 0) for c in pension_cases)
            if total_pension > 0:
                summary_data['Monto Mensual Pensión'] = [
                    f"${total_pension:,.2f}",
                    "-",
                    "-",
                    f"${total_pension:,.2f}"
                ]
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Resumen General', index=False)
            self._format_sheet(writer.book['Resumen General'])
            
            # 2. PENSIÓN ALIMENTARIA
            if pension_cases:
                pension_df = self._cases_to_dataframe(pension_cases)
                pension_df.to_excel(writer, sheet_name='Pensión Alimentaria', index=False)
                self._format_sheet(writer.book['Pensión Alimentaria'])
                
                # Pension statistics
                pension_stats = self._get_category_breakdown(pension_cases)
                pension_stats_df = pd.DataFrame(list(pension_stats.items()), 
                                               columns=['Tipo', 'Cantidad'])
                pension_stats_df = pension_stats_df.sort_values('Cantidad', ascending=False)
                pension_stats_df.to_excel(writer, sheet_name='Estadísticas Pensión', index=False)
                self._format_sheet(writer.book['Estadísticas Pensión'])
            
            # 3. TRÁNSITO
            if transito_cases:
                transito_df = self._cases_to_dataframe(transito_cases)
                transito_df.to_excel(writer, sheet_name='Tránsito', index=False)
                self._format_sheet(writer.book['Tránsito'])
                
                # Transit statistics
                transito_stats = self._get_category_breakdown(transito_cases)
                transito_stats_df = pd.DataFrame(list(transito_stats.items()), 
                                                columns=['Tipo', 'Cantidad'])
                transito_stats_df = transito_stats_df.sort_values('Cantidad', ascending=False)
                transito_stats_df.to_excel(writer, sheet_name='Estadísticas Tránsito', index=False)
                self._format_sheet(writer.book['Estadísticas Tránsito'])
            
            # 4. OTROS CASOS
            if otros_cases:
                otros_df = self._cases_to_dataframe(otros_cases)
                otros_df.to_excel(writer, sheet_name='Otros Casos', index=False)
                self._format_sheet(writer.book['Otros Casos'])
                
                # Other cases statistics
                otros_stats = self._get_category_breakdown(otros_cases)
                otros_stats_df = pd.DataFrame(list(otros_stats.items()), 
                                              columns=['Tipo', 'Cantidad'])
                otros_stats_df = otros_stats_df.sort_values('Cantidad', ascending=False)
                otros_stats_df.to_excel(writer, sheet_name='Estadísticas Otros', index=False)
                self._format_sheet(writer.book['Estadísticas Otros'])
            
            # 5. TIPOS DE RESOLUCIÓN
            resolution_stats = self._get_resolution_statistics(cases)
            resolution_df = pd.DataFrame(list(resolution_stats.items()), 
                                        columns=['Tipo de Resolución', 'Cantidad'])
            resolution_df = resolution_df.sort_values('Cantidad', ascending=False)
            resolution_df.to_excel(writer, sheet_name='Tipos de Resolución', index=False)
            self._format_sheet(writer.book['Tipos de Resolución'])
        
        return filepath
    
    def _cases_to_dataframe(self, cases):
        """Convert cases to pandas DataFrame"""
        data = []
        for case in cases:
            data.append({
                'Número Carpeta': case.numero_carpeta,
                'Categoría': case.categoria,
                'Víctima': case.victima,
                'Investigado': case.investigado,
                'Fiscal': case.fiscal_asignado,
                'Etapa': case.etapa_procesal,
                'Estado': case.estado_actual,
                'Fecha Denuncia': case.fecha_denuncia,
                'Monto Pensión': case.monto_pension if hasattr(case, 'monto_pension') else None,
                'Observaciones': case.estado_actual
            })
        return pd.DataFrame(data)
    
    def _get_category_breakdown(self, cases):
        """Get breakdown by category"""
        breakdown = {}
        for case in cases:
            cat = case.categoria if case.categoria else 'Sin categoría'
            breakdown[cat] = breakdown.get(cat, 0) + 1
        return breakdown
    
    def _get_resolution_statistics(self, cases):
        """Get resolution type statistics"""
        resolution_types = {
            'Conciliaciones': 0,
            'Condenas': 0,
            'Acuerdos': 0,
            'No acuerdos': 0,
            'Desistimientos': 0,
            'Archivos': 0,
            'En proceso': 0
        }
        
        for case in cases:
            cat_lower = case.categoria.lower() if case.categoria else ''
            if 'conciliación' in cat_lower or 'conciliacion' in cat_lower:
                resolution_types['Conciliaciones'] += 1
            elif 'condena' in cat_lower:
                resolution_types['Condenas'] += 1
            elif 'acuerdo' in cat_lower and 'no acuerdo' not in cat_lower:
                resolution_types['Acuerdos'] += 1
            elif 'no acuerdo' in cat_lower:
                resolution_types['No acuerdos'] += 1
            elif 'desistimiento' in cat_lower:
                resolution_types['Desistimientos'] += 1
            elif 'archivo' in cat_lower:
                resolution_types['Archivos'] += 1
            else:
                resolution_types['En proceso'] += 1
        
        return resolution_types
    
    def _format_sheet(self, worksheet):
        """Apply formatting to worksheet"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Format header
        header_fill = PatternFill(start_color='3b82f6', end_color='3b82f6', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column width
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
